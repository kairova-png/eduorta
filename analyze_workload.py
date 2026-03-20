#!/usr/bin/env python3
from openpyxl import load_workbook

file_path = '/home/leo/Desktop/AUTO-SCHEDULE/ALL_workload_detailed.xlsx'
wb = load_workbook(file_path, read_only=True)
ws = wb['По дисциплинам']

teachers = set()
groups = set()
depts = set()

for row in range(2, ws.max_row + 1):
    d = ws.cell(row=row, column=1).value
    t = ws.cell(row=row, column=2).value
    g = ws.cell(row=row, column=5).value
    if d: depts.add(d)
    if t: teachers.add(t)
    if g: groups.add(g)

print(f'Записей: {ws.max_row - 1}')
print(f'Отделений: {len(depts)}')
for d in sorted(depts): print(f'  {d}')
print(f'Групп: {len(groups)}')
for g in sorted(groups): print(f'  {g}')
print(f'Преподавателей: {len(teachers)}')
for t in sorted(teachers): print(f'  {t}')
wb.close()
