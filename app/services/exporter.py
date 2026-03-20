import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from config import Config


class ExcelExporter:
    """Service for exporting data to Excel"""

    def __init__(self, exports_dir):
        self.exports_dir = exports_dir
        os.makedirs(exports_dir, exist_ok=True)

    def export_schedule_week(self, group, schedules, week_start, week_end):
        """Export weekly schedule for a group"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Status colors
        status_fills = {
            'planned': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
            'done': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
            'substitution': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
            'cancelled': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
            'completed': PatternFill(start_color='E2D5F1', end_color='E2D5F1', fill_type='solid'),
        }

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f'APEC Petrotechnic - Расписание'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:G2')
        ws['A2'] = f'Группа: {group.name}    Неделя: {week_start.strftime("%d.%m.%Y")} - {week_end.strftime("%d.%m.%Y")}'
        ws['A2'].font = title_font
        ws['A2'].alignment = center_align

        # Headers
        headers = ['Пара', 'Время', 'Пн', 'Вт', 'Ср', 'Чт', 'Пт']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = center_align
            cell.border = cell_border

        # Pair times
        pair_times = Config.PAIR_TIMES

        # Build schedule grid
        schedule_by_day_pair = {}
        for s in schedules:
            key = (s.date.weekday(), s.pair_number)
            schedule_by_day_pair[key] = s

        # Fill schedule data
        row = 5
        for pair_num in range(1, 8):
            times = pair_times.get(pair_num, {})
            time_str = f"{times.get('start', '')}-{times.get('end', '')}"

            # Pair number
            cell = ws.cell(row=row, column=1, value=pair_num)
            cell.alignment = center_align
            cell.border = cell_border

            # Time
            cell = ws.cell(row=row, column=2, value=time_str)
            cell.alignment = center_align
            cell.border = cell_border

            # Days Mon-Fri (weekday 0-4)
            for day_idx in range(5):
                col = day_idx + 3
                s = schedule_by_day_pair.get((day_idx, pair_num))

                if s:
                    content = f"{s.workload.discipline}\n{s.workload.teacher.full_name}\nауд. {s.room.number if s.room else '-'}"
                    cell = ws.cell(row=row, column=col, value=content)
                    cell.fill = status_fills.get(s.status, status_fills['planned'])
                else:
                    cell = ws.cell(row=row, column=col, value='')

                cell.alignment = center_align
                cell.border = cell_border

            row += 1

            # Add lunch break after pair 3
            if pair_num == 3:
                ws.cell(row=row, column=1, value='').border = cell_border
                cell = ws.cell(row=row, column=2, value='ОБЕД')
                cell.alignment = center_align
                cell.border = cell_border
                cell.font = Font(bold=True)
                for col in range(3, 8):
                    ws.cell(row=row, column=col, value='').border = cell_border
                row += 1

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        for col in range(3, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Row heights
        for r in range(5, row):
            ws.row_dimensions[r].height = 50

        # Save file
        filename = f"schedule_{group.name}_{week_start.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        wb.save(filepath)
        return filepath

    def export_hours_report(self, group, workloads, semester):
        """Export hours tracking report for a group"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Учет часов"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f'Учет часов - {group.name}, {semester} семестр'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        # Headers
        headers = ['Дисциплина', 'Тип', 'Преподаватель', 'План', 'Факт', 'Остаток', '%']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = title_font
            cell.alignment = center_align
            cell.border = cell_border

        # Data rows
        row = 4
        total_plan = 0
        total_fact = 0

        for wl in workloads:
            ws.cell(row=row, column=1, value=wl.discipline).border = cell_border
            # Lesson type labels: T=theory, P=practice, K=consultation, E=exam
            type_labels = {'theory': 'Т', 'practice': 'П', 'consultation': 'К', 'exam': 'Э'}
            ws.cell(row=row, column=2, value=type_labels.get(wl.lesson_type, 'Т')).border = cell_border
            ws.cell(row=row, column=3, value=wl.teacher.full_name if wl.teacher else '-').border = cell_border
            ws.cell(row=row, column=4, value=wl.total_hours).border = cell_border
            ws.cell(row=row, column=5, value=wl.hours_completed).border = cell_border
            ws.cell(row=row, column=6, value=wl.hours_remaining).border = cell_border
            ws.cell(row=row, column=7, value=f"{wl.progress_percent}%").border = cell_border

            for col in range(1, 8):
                ws.cell(row=row, column=col).alignment = center_align

            total_plan += wl.total_hours
            total_fact += wl.hours_completed
            row += 1

        # Total row
        ws.cell(row=row, column=1, value='ИТОГО').font = title_font
        ws.cell(row=row, column=4, value=total_plan).font = title_font
        ws.cell(row=row, column=5, value=total_fact).font = title_font
        ws.cell(row=row, column=6, value=total_plan - total_fact).font = title_font
        progress = int((total_fact / total_plan * 100)) if total_plan > 0 else 0
        ws.cell(row=row, column=7, value=f"{progress}%").font = title_font

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = cell_border
            ws.cell(row=row, column=col).alignment = center_align

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10

        # Save
        filename = f"hours_{group.name}_sem{semester}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        wb.save(filepath)
        return filepath

    def export_teacher_timesheet(self, teacher, schedules, month, year):
        """Export teacher timesheet for a month"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Табель"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Month names
        months_ru = {
            1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
            5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
            9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
        }

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Табель: {teacher.full_name}, {months_ru[month]} {year}'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        # Headers
        headers = ['Группа', 'Дисциплина', 'Тип', 'Основные', 'Замены', 'Итого']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = title_font
            cell.alignment = center_align
            cell.border = cell_border

        # Group schedules by workload
        workload_hours = {}
        substitution_hours = 0

        for s in schedules:
            if s.status in ['done', 'completed']:
                wl_id = s.workload_id
                if wl_id not in workload_hours:
                    workload_hours[wl_id] = {
                        'workload': s.workload,
                        'hours': 0
                    }
                workload_hours[wl_id]['hours'] += 2
            elif s.status == 'substitution' and s.substitute_teacher_id == teacher.id:
                substitution_hours += 2

        # Data rows
        row = 4
        total_main = 0
        total_sub = 0

        for wl_id, data in workload_hours.items():
            wl = data['workload']
            hours = data['hours']

            ws.cell(row=row, column=1, value=wl.group.name if wl.group else '-').border = cell_border
            ws.cell(row=row, column=2, value=wl.discipline).border = cell_border
            type_labels = {'theory': 'Т', 'practice': 'П', 'consultation': 'К', 'exam': 'Э'}
            ws.cell(row=row, column=3, value=type_labels.get(wl.lesson_type, 'Т')).border = cell_border
            ws.cell(row=row, column=4, value=hours).border = cell_border
            ws.cell(row=row, column=5, value=0).border = cell_border
            ws.cell(row=row, column=6, value=hours).border = cell_border

            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = center_align

            total_main += hours
            row += 1

        # Substitutions row if any
        if substitution_hours > 0:
            ws.cell(row=row, column=1, value='-').border = cell_border
            ws.cell(row=row, column=2, value='Замены (вакант)').border = cell_border
            ws.cell(row=row, column=3, value='-').border = cell_border
            ws.cell(row=row, column=4, value=0).border = cell_border
            ws.cell(row=row, column=5, value=substitution_hours).border = cell_border
            ws.cell(row=row, column=6, value=substitution_hours).border = cell_border
            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = center_align
            total_sub = substitution_hours
            row += 1

        # Total row
        ws.cell(row=row, column=1, value='').border = cell_border
        ws.cell(row=row, column=2, value='ИТОГО за месяц').font = title_font
        ws.cell(row=row, column=2).border = cell_border
        ws.cell(row=row, column=3, value='').border = cell_border
        ws.cell(row=row, column=4, value=total_main).font = title_font
        ws.cell(row=row, column=4).border = cell_border
        ws.cell(row=row, column=5, value=total_sub).font = title_font
        ws.cell(row=row, column=5).border = cell_border
        ws.cell(row=row, column=6, value=total_main + total_sub).font = title_font
        ws.cell(row=row, column=6).border = cell_border

        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = center_align

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10

        # Save
        filename = f"timesheet_{teacher.id}_{year}_{month:02d}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        wb.save(filepath)
        return filepath

    def export_teacher_timesheet_ministry(self, teacher, month, year):
        """Export teacher timesheet using ministry approved template"""
        from datetime import datetime
        from calendar import monthrange
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ведомость"

        # Styles
        header_font = Font(bold=True, size=12, name='Times New Roman')
        title_font = Font(bold=True, size=10, name='Times New Roman')
        regular_font = Font(size=10, name='Times New Roman')
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Get academic year and semester
        from app.models import AcademicYear
        current_academic_year = AcademicYear.query.filter_by(is_current=True).first()
        semester = 1 if month <= 6 else 2  # Simple logic for semester
        
        # Month names 
        months_ru = {
            1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
            5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
            9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
        }

        # Header section
        ws.merge_cells('A1:L1')
        ws['A1'] = 'Министерство Просвещения Республики Казахстан'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:L2')
        ws['A2'] = 'ТОО «Высший колледж APEC PetroTechnic»'
        ws['A2'].font = header_font
        ws['A2'].alignment = center_align

        year_text = f"{current_academic_year.name}" if current_academic_year else f"{year}-{year+1}"
        ws.merge_cells('A3:L3')
        ws['A3'] = f'Ведомость учета учебного времени педагога за {semester}-ый семестр {year_text} учебного года'
        ws['A3'].font = header_font
        ws['A3'].alignment = center_align

        # Teacher name line
        ws.merge_cells('A5:L5')
        ws['A5'] = f'Фамилия, имя, отчество (при его наличии) педагога (полностью) {teacher.full_name}'
        ws['A5'].font = regular_font
        ws['A5'].alignment = left_align

        # Get teacher workloads and schedule data
        from app.models import Workload, Schedule
        teacher_workloads = Workload.query.filter_by(teacher_id=teacher.id).all()
        
        # Get date range for month
        _, last_day = monthrange(year, month)
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, last_day).date()
        
        # Get completed schedules for the month
        completed_schedules = Schedule.query.join(Workload).filter(
            Workload.teacher_id == teacher.id,
            Schedule.date >= start_date,
            Schedule.date <= end_date,
            Schedule.status.in_(['done', 'completed'])
        ).all()
        
        # Get substitution schedules
        substitution_schedules = Schedule.query.filter(
            Schedule.substitute_teacher_id == teacher.id,
            Schedule.date >= start_date,
            Schedule.date <= end_date,
            Schedule.status == 'substitution'
        ).all()

        # Main workload table
        row_start = 7
        
        # Table header - first row
        ws.cell(row=row_start, column=1, value='Группы/Месяцы').font = title_font
        ws.cell(row=row_start, column=1).alignment = center_align
        ws.cell(row=row_start, column=1).border = cell_border
        
        ws.merge_cells(f'B{row_start}:K{row_start}')
        ws.cell(row=row_start, column=2, value='По нагрузке').font = title_font
        ws.cell(row=row_start, column=2).alignment = center_align
        ws.cell(row=row_start, column=2).border = cell_border
        
        # Add borders to all merged cells in header
        for col in range(3, 12):
            ws.cell(row=row_start, column=col).border = cell_border
        
        ws.cell(row=row_start, column=12, value='Всего').font = title_font
        ws.cell(row=row_start, column=12).alignment = center_align
        ws.cell(row=row_start, column=12).border = cell_border

        # Add workload columns (groups and disciplines)
        col_start = 2
        group_disciplines = {}
        for wl in teacher_workloads:
            if not wl.discipline.startswith('ЗАМЕНА:'):  # Exclude substitution workloads
                key = f"{wl.group.name if wl.group else 'Без группы'}"
                if key not in group_disciplines:
                    group_disciplines[key] = []
                group_disciplines[key].append(wl)
        
        # Add group/discipline headers
        col = col_start
        for group_name, workloads in group_disciplines.items():
            for wl in workloads:
                if col <= 11:  # Limit columns
                    ws.cell(row=row_start+1, column=col, value=f"{group_name}\n{wl.discipline}").font = regular_font
                    ws.cell(row=row_start+1, column=col).alignment = center_align
                    ws.cell(row=row_start+1, column=col).border = cell_border
                    col += 1
        
        # Fill empty header columns with borders
        while col <= 11:
            ws.cell(row=row_start+1, column=col, value="").border = cell_border
            col += 1
        
        # Months section
        current_row = row_start + 2
        
        # Add current month row
        ws.cell(row=current_row, column=1, value=months_ru[month]).font = regular_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = cell_border
        
        # Calculate hours for each workload in this month
        total_month_hours = 0
        col = col_start
        for group_name, workloads in group_disciplines.items():
            for wl in workloads:
                if col <= 11:
                    # Count hours for this workload in the month
                    wl_hours = 0
                    for s in completed_schedules:
                        if s.workload_id == wl.id:
                            wl_hours += 2
                    
                    ws.cell(row=current_row, column=col, value=wl_hours if wl_hours > 0 else '').font = regular_font
                    ws.cell(row=current_row, column=col).alignment = center_align
                    ws.cell(row=current_row, column=col).border = cell_border
                    total_month_hours += wl_hours
                    col += 1
        
        # Fill empty month data columns with borders
        while col <= 11:
            ws.cell(row=current_row, column=col, value="").border = cell_border
            col += 1
        
        # Month total
        ws.cell(row=current_row, column=12, value=total_month_hours).font = title_font
        ws.cell(row=current_row, column=12).alignment = center_align
        ws.cell(row=current_row, column=12).border = cell_border
        
        # Summary rows
        current_row += 2

        # Calculate exam and consultation hours from workloads
        exam_hours = 0
        consultation_hours = 0
        for wl in teacher_workloads:
            if wl.lesson_type == 'exam':
                # Count completed exam schedules for this workload in the month
                for s in completed_schedules:
                    if s.workload_id == wl.id:
                        exam_hours += 2
            elif wl.lesson_type == 'consultation':
                # Count completed consultation schedules
                for s in completed_schedules:
                    if s.workload_id == wl.id:
                        consultation_hours += 2

        # Экзамены
        ws.cell(row=current_row, column=1, value='Экзамены').font = regular_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = cell_border
        # Fill empty columns with borders
        for col in range(2, 12):
            ws.cell(row=current_row, column=col, value="").border = cell_border
        ws.cell(row=current_row, column=12, value=exam_hours).font = regular_font
        ws.cell(row=current_row, column=12).alignment = center_align
        ws.cell(row=current_row, column=12).border = cell_border
        current_row += 1

        # Консультации
        ws.cell(row=current_row, column=1, value='Консультации').font = regular_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = cell_border
        # Fill empty columns with borders
        for col in range(2, 12):
            ws.cell(row=current_row, column=col, value="").border = cell_border
        ws.cell(row=current_row, column=12, value=consultation_hours).font = regular_font
        ws.cell(row=current_row, column=12).alignment = center_align
        ws.cell(row=current_row, column=12).border = cell_border
        current_row += 1
        
        # Всего запланировано - считаем по отображаемым workload
        total_planned = 0
        for group_name, workloads in group_disciplines.items():
            for wl in workloads:
                total_planned += wl.total_hours
        ws.cell(row=current_row, column=1, value='Всего запланировано, часов').font = title_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = cell_border
        # Fill empty columns with borders
        for col in range(2, 12):
            ws.cell(row=current_row, column=col, value="").border = cell_border
        ws.cell(row=current_row, column=12, value=total_planned).font = title_font
        ws.cell(row=current_row, column=12).alignment = center_align
        ws.cell(row=current_row, column=12).border = cell_border
        current_row += 1
        
        # Фактически выполнено (без замен)
        ws.cell(row=current_row, column=1, value='Фактически выполнено, часов').font = title_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = cell_border
        # Fill empty columns with borders
        for col in range(2, 12):
            ws.cell(row=current_row, column=col, value="").border = cell_border
        ws.cell(row=current_row, column=12, value=total_month_hours).font = title_font
        ws.cell(row=current_row, column=12).alignment = center_align
        ws.cell(row=current_row, column=12).border = cell_border
        current_row += 2
        
        # Summary text
        substitution_hours = len(substitution_schedules) * 2
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = f'Всего часов за {months_ru[month]} месяц {total_month_hours} по нагрузке - ; Замена часов-{substitution_hours};'
        ws[f'A{current_row}'].font = regular_font
        ws[f'A{current_row}'].alignment = left_align
        current_row += 2

        # Substitutions table
        if substitution_schedules:
            ws.merge_cells(f'A{current_row}:L{current_row}')
            ws[f'A{current_row}'] = 'Замена часов (только по дисциплине/модулю специальности)'
            ws[f'A{current_row}'].font = title_font
            ws[f'A{current_row}'].alignment = center_align
            current_row += 1
            
            # Substitution table header
            ws.cell(row=current_row, column=1, value='Группы/Месяц').font = title_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = cell_border
            
            ws.merge_cells(f'B{current_row}:L{current_row}')
            ws.cell(row=current_row, column=2, value='Замены').font = title_font
            ws.cell(row=current_row, column=2).alignment = center_align
            ws.cell(row=current_row, column=2).border = cell_border
            
            # Add borders to merged cells
            for col in range(3, 13):
                ws.cell(row=current_row, column=col).border = cell_border
            current_row += 1
            
            # Month row for substitutions
            ws.cell(row=current_row, column=1, value=months_ru[month]).font = regular_font
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=1).border = cell_border
            
            # Group substitution data
            sub_text = ""
            for s in substitution_schedules:
                original_teacher = s.workload.teacher.full_name
                group_name = s.workload.group.name if s.workload.group else 'Без группы'
                discipline = s.workload.discipline
                sub_text += f"{group_name} {discipline} (замена {original_teacher}) 2ч; "
            
            ws.merge_cells(f'B{current_row}:L{current_row}')
            ws.cell(row=current_row, column=2, value=sub_text).font = regular_font
            ws.cell(row=current_row, column=2).alignment = left_align
            ws.cell(row=current_row, column=2).border = cell_border
            
            # Add borders to merged cells
            for col in range(3, 13):
                ws.cell(row=current_row, column=col).border = cell_border
            current_row += 2

        # Signature section
        current_row += 2
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = f'Подпись преподавателя : ________________________                            Дата________________        {year}г.'
        ws[f'A{current_row}'].font = regular_font
        current_row += 4
        
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = 'Заместитель директора по академической работе                            __________________                            _________'
        ws[f'A{current_row}'].font = regular_font
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = '                                                                                                                                                                (подпись)'
        ws[f'A{current_row}'].font = regular_font

        # Column widths
        for col in range(1, 13):
            ws.column_dimensions[chr(64 + col)].width = 12
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25

        # Save
        filename = f"ministry_timesheet_{teacher.id}_{year}_{month:02d}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        wb.save(filepath)
        return filepath

    def export_period_timesheet(self, period, schedules, academic_year):
        """Export timesheet for a specific period/week"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Табель за период"

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f'APEC Petrotechnic - Табель учета часов'
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:E2')
        ws['A2'] = f'Период: Неделя {period.week_number} ({period.start_date.strftime("%d.%m.%Y")} - {period.end_date.strftime("%d.%m.%Y")})'
        ws['A2'].font = title_font
        ws['A2'].alignment = center_align

        ws.merge_cells('A3:E3')
        ws['A3'] = f'Учебный год: {academic_year.name}'
        ws['A3'].alignment = center_align

        # Headers
        headers = ['№', 'Преподаватель', 'Дисциплины', 'Пар', 'Часов']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = title_font
            cell.alignment = center_align
            cell.border = cell_border

        # Group schedules by teacher
        from app.models import Teacher
        teacher_hours = {}

        for s in schedules:
            teacher = s.workload.teacher

            # Check if it's a substitution
            if s.status == 'substitution' and s.substitute_teacher_id:
                teacher = Teacher.query.get(s.substitute_teacher_id)

            if teacher.id not in teacher_hours:
                teacher_hours[teacher.id] = {
                    'teacher': teacher,
                    'pairs': 0,
                    'disciplines': {}
                }

            teacher_hours[teacher.id]['pairs'] += 1
            discipline = s.workload.discipline
            if discipline not in teacher_hours[teacher.id]['disciplines']:
                teacher_hours[teacher.id]['disciplines'][discipline] = 0
            teacher_hours[teacher.id]['disciplines'][discipline] += 1

        # Sort by teacher name
        teacher_data = sorted(teacher_hours.values(), key=lambda x: x['teacher'].full_name)

        # Data rows
        row = 6
        total_pairs = 0

        for idx, data in enumerate(teacher_data, 1):
            # Row number
            cell = ws.cell(row=row, column=1, value=idx)
            cell.alignment = center_align
            cell.border = cell_border

            # Teacher name
            cell = ws.cell(row=row, column=2, value=data['teacher'].full_name)
            cell.alignment = Alignment(vertical='center')
            cell.border = cell_border

            # Disciplines with counts
            disciplines_str = ', '.join([f"{d}: {c}" for d, c in data['disciplines'].items()])
            cell = ws.cell(row=row, column=3, value=disciplines_str)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = cell_border

            # Pairs count
            cell = ws.cell(row=row, column=4, value=data['pairs'])
            cell.alignment = center_align
            cell.border = cell_border

            # Hours (pairs * 2)
            cell = ws.cell(row=row, column=5, value=data['pairs'] * 2)
            cell.alignment = center_align
            cell.border = cell_border

            total_pairs += data['pairs']
            row += 1

        # Total row
        ws.cell(row=row, column=1, value='').border = cell_border
        cell = ws.cell(row=row, column=2, value='ИТОГО')
        cell.font = title_font
        cell.alignment = center_align
        cell.border = cell_border
        ws.cell(row=row, column=3, value='').border = cell_border
        cell = ws.cell(row=row, column=4, value=total_pairs)
        cell.font = title_font
        cell.alignment = center_align
        cell.border = cell_border
        cell = ws.cell(row=row, column=5, value=total_pairs * 2)
        cell.font = title_font
        cell.alignment = center_align
        cell.border = cell_border

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10

        # Save
        filename = f"period_timesheet_week{period.week_number}_{period.start_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        wb.save(filepath)
        return filepath
