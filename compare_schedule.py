#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Compare schedule from database with Excel file
Period: 12.01.2026 - 16.01.2026
"""

import sys
import os
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from datetime import date, timedelta
from app import create_app, db
from app.models import Schedule, Workload, Group, Teacher, Room
import openpyxl
from collections import defaultdict

app = create_app()

EXCEL_FILE = r"C:\Users\APEC-mono49\Desktop\Apec\college_schedule_backup_20251212\college_schedule\Копия 1 курс Расписание 2 семестр.xlsx"
SHEET_NAME = "II cем 12.01.-16.01.25"
START_DATE = date(2026, 1, 12)
END_DATE = date(2026, 1, 16)

WEEKDAYS = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
WEEKDAYS_SHORT = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']

# Output file
OUTPUT_FILE = "comparison_result.txt"

def log(msg):
    """Write to both console and file"""
    print(msg)
    with open(OUTPUT_FILE, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')

def get_db_schedule():
    """Get schedule from database for the period"""
    with app.app_context():
        schedules = Schedule.query.filter(
            Schedule.date >= START_DATE,
            Schedule.date <= END_DATE
        ).order_by(Schedule.date, Schedule.pair_number).all()

        db_data = defaultdict(lambda: defaultdict(dict))

        for s in schedules:
            if s.workload and s.workload.group:
                group_name = s.workload.group.name
                day = s.date
                pair_num = s.pair_number

                teacher_name = s.workload.teacher.full_name if s.workload.teacher else 'N/A'
                discipline = s.workload.discipline
                room = s.room.number if s.room else 'N/A'

                db_data[group_name][day][pair_num] = {
                    'discipline': discipline,
                    'teacher': teacher_name,
                    'room': room
                }

        return db_data

def read_excel_schedule():
    """Read schedule from Excel file - improved parsing"""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    excel_data = defaultdict(lambda: defaultdict(dict))

    # Row 5 has group names at columns 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34
    # Discipline data is in the same columns, teacher in col+1
    group_columns = {}

    # Extract groups from row 5
    for col in range(4, 40, 2):  # Even columns from 4 to 38
        cell_value = ws.cell(row=5, column=col).value
        if cell_value:
            group_name = str(cell_value).strip()
            # Only keep actual group names (contain pattern like X-25 or X-24)
            if '-25' in group_name or '-24' in group_name:
                group_columns[col] = group_name

    log(f"Found {len(group_columns)} groups in Excel: {list(group_columns.values())}")

    # Day mapping
    day_map = {
        'ПОНЕДЕЛЬНИК': date(2026, 1, 12),
        'ВТОРНИК': date(2026, 1, 13),
        'СРЕДА': date(2026, 1, 14),
        'ЧЕТВЕРГ': date(2026, 1, 15),
        'ПЯТНИЦА': date(2026, 1, 16),
    }

    # Parse data starting from row 6
    # Structure: Each day has pairs 1-4, each pair has 2 rows (discipline, teacher)
    current_day = None
    current_pair = None

    for row_idx in range(6, ws.max_row + 1):
        # Column 1 has day names
        col1 = ws.cell(row=row_idx, column=1).value
        if col1:
            col1_str = str(col1).strip().upper()
            for day_name, day_date in day_map.items():
                if day_name in col1_str:
                    current_day = day_date
                    break

        # Column 3 has pair numbers (I, II, III, IV or 1, 2, 3, 4)
        col3 = ws.cell(row=row_idx, column=3).value
        if col3:
            col3_str = str(col3).strip()
            # Roman to Arabic
            roman_map = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6}
            if col3_str in roman_map:
                current_pair = roman_map[col3_str]
            elif col3_str.isdigit():
                current_pair = int(col3_str)

        if not current_day or not current_pair:
            continue

        # Get disciplines for each group
        for col, group_name in group_columns.items():
            cell_value = ws.cell(row=row_idx, column=col).value
            if cell_value:
                discipline = str(cell_value).strip()
                # Skip empty or invalid
                if not discipline or discipline == '-' or len(discipline) < 3:
                    continue
                # Skip if it looks like a teacher name (contains only initials)
                if '.' in discipline and len(discipline) < 20:
                    continue

                # Get teacher from next row or next column
                teacher = ''
                teacher_cell = ws.cell(row=row_idx + 1, column=col).value
                if teacher_cell:
                    teacher = str(teacher_cell).strip()

                excel_data[group_name][current_day][current_pair] = {
                    'discipline': discipline,
                    'teacher': teacher,
                    'room': ''
                }

    return excel_data

def normalize_name(name):
    """Normalize discipline name for comparison"""
    if not name:
        return ''
    name = str(name).strip().lower()
    name = name.replace('ё', 'е').replace('  ', ' ')
    # Remove slashes and take first part for compound disciplines
    if '/' in name:
        name = name.split('/')[0].strip()
    return name

def compare_schedules(db_data, excel_data):
    """Compare two schedules"""
    differences = []
    matches = []

    # Only compare groups that exist in BOTH
    common_groups = set(db_data.keys()) & set(excel_data.keys())
    log(f"\nCommon groups: {sorted(common_groups)}")

    only_db = set(db_data.keys()) - set(excel_data.keys())
    only_excel = set(excel_data.keys()) - set(db_data.keys())

    if only_db:
        log(f"Groups only in DB: {sorted(only_db)}")
    if only_excel:
        log(f"Groups only in Excel: {sorted(only_excel)}")

    for group in sorted(common_groups):
        db_group = db_data.get(group, {})
        excel_group = excel_data.get(group, {})

        all_dates = set(db_group.keys()) | set(excel_group.keys())

        for day in sorted(all_dates):
            db_day = db_group.get(day, {})
            excel_day = excel_group.get(day, {})

            all_pairs = set(db_day.keys()) | set(excel_day.keys())

            for pair in sorted(all_pairs):
                db_pair = db_day.get(pair, {})
                excel_pair = excel_day.get(pair, {})

                db_disc = normalize_name(db_pair.get('discipline', ''))
                excel_disc = normalize_name(excel_pair.get('discipline', ''))

                weekday = WEEKDAYS_SHORT[day.weekday()] if isinstance(day, date) else '?'

                if db_disc == excel_disc and db_disc:
                    matches.append({
                        'group': group,
                        'date': day,
                        'weekday': weekday,
                        'pair': pair,
                        'discipline': db_pair.get('discipline', ''),
                    })
                elif db_disc != excel_disc:
                    differences.append({
                        'group': group,
                        'date': day,
                        'weekday': weekday,
                        'pair': pair,
                        'db_discipline': db_pair.get('discipline', '-') if db_pair else '-',
                        'db_teacher': db_pair.get('teacher', '-') if db_pair else '-',
                        'excel_discipline': excel_pair.get('discipline', '-') if excel_pair else '-',
                        'excel_teacher': excel_pair.get('teacher', '-') if excel_pair else '-',
                    })

    return differences, matches

def main():
    # Clear output file
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('')

    log("=" * 80)
    log("COMPARISON: Database Schedule vs Excel")
    log(f"Period: {START_DATE} - {END_DATE}")
    log("=" * 80)

    log("\n1. Reading database schedule...")
    db_data = get_db_schedule()
    log(f"   Found {len(db_data)} groups in database")

    for group, dates in sorted(db_data.items()):
        total_pairs = sum(len(pairs) for pairs in dates.values())
        log(f"   - {group}: {total_pairs} pairs")

    log("\n2. Reading Excel schedule...")
    excel_data = read_excel_schedule()
    log(f"   Found {len(excel_data)} groups in Excel")

    for group, dates in sorted(excel_data.items()):
        total_pairs = sum(len(pairs) for pairs in dates.values())
        log(f"   - {group}: {total_pairs} pairs")

    log("\n3. Comparing schedules...")
    differences, matches = compare_schedules(db_data, excel_data)

    log(f"\n" + "=" * 80)
    log(f"RESULTS: {len(matches)} matches, {len(differences)} differences")
    log("=" * 80)

    if differences:
        log(f"\n### DIFFERENCES ({len(differences)}) ###\n")

        by_group = defaultdict(list)
        for diff in differences:
            by_group[diff['group']].append(diff)

        for group in sorted(by_group.keys()):
            log(f"\n{group}:")
            log("-" * 60)
            for diff in sorted(by_group[group], key=lambda x: (x['date'], x['pair'])):
                day_str = diff['date'].strftime('%d.%m') if isinstance(diff['date'], date) else str(diff['date'])
                log(f"  {diff['weekday']} {day_str}, Para {diff['pair']}:")
                log(f"    DB:    {diff['db_discipline']}")
                log(f"    Excel: {diff['excel_discipline']}")

    # Show full DB schedule for 1-25 groups only
    log("\n" + "=" * 80)
    log("FULL DATABASE SCHEDULE (1 course groups):")
    log("=" * 80)

    for group in sorted(db_data.keys()):
        if '1-25' not in group:
            continue
        log(f"\n### {group} ###")
        for day in sorted(db_data[group].keys()):
            weekday = WEEKDAYS_SHORT[day.weekday()]
            log(f"  {weekday} {day.strftime('%d.%m')}:")
            for pair in sorted(db_data[group][day].keys()):
                info = db_data[group][day][pair]
                log(f"    {pair}. {info['discipline']} - {info['teacher']} [{info['room']}]")

    log(f"\n\nResults saved to: {OUTPUT_FILE}")

if __name__ == '__main__':
    main()
