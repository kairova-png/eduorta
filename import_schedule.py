# -*- coding: utf-8 -*-
"""
Import schedule from Excel files for courses 2 and 4.
Creates Schedule entries and WorkloadTemplate entries.
"""

import openpyxl
import sqlite3
import json
from datetime import datetime, timedelta

DB_PATH = 'instance/college.db'

DATE_START = datetime(2026, 2, 2)
DATE_END = datetime(2026, 3, 6)

EXCEL_FILES = [
    {
        'path': 'exports/Расписание 4 курс 1 смена.xlsx',
        'sheet_index': 0,
        'sheet_index_week1': 1,
        'enrollment_year': 2022,
        'semester': 8,
        'header_row': 6,
        'data_start_row': 7,
        'pair_time_map': {'I': 1, 'II': 2, 'III': 3, 'IV': 4}
    },
    {
        'path': 'exports/Расписание 2 курс 2 смена.xlsx',
        'sheet_index': 0,
        'sheet_index_week1': 1,
        'enrollment_year': 2024,
        'semester': 4,
        'header_row': 6,
        'data_start_row': 7,
        'pair_time_map': {'I': 4, 'II': 5, 'III': 6}
    }
]

DAY_NAMES = {
    'ПОНЕДЕЛЬНИК': 0, 'ВТОРНИК': 1, 'СРЕДА': 2,
    'ЧЕТВЕРГ': 3, 'ПЯТНИЦА': 4, 'СУББОТА': 5,
}

ROOM_SPECIAL_MAP = {
    'маст': 'Мастерская', 'мастерская': 'Мастерская',
    'с/з': 'С/з', 'с/З': 'С/з', 'С/З': 'С/з', 'С/з': 'С/з',
    'спортзал': 'Спортзал',
}

TEACHER_NORMALIZE = {
    'Айбарұлы Ғ.': 'Айбарұлы Г.',
    'Темирбулатов Р.': 'Темирбулатов Р.М.',
    'Тастанбеков С..': 'Тастанбеков С.',
    'Измуканова А.': 'Измуханова А.',
}


def load_db_data(conn):
    cur = conn.cursor()

    cur.execute('SELECT id, name, enrollment_year FROM groups')
    groups = {r[1]: {'id': r[0], 'enrollment_year': r[2]} for r in cur.fetchall()}

    cur.execute('SELECT id, full_name FROM teachers')
    teachers = {r[1]: r[0] for r in cur.fetchall()}

    cur.execute('SELECT id, number FROM rooms')
    rooms = {r[1]: r[0] for r in cur.fetchall()}

    cur.execute('''SELECT w.id, w.group_id, w.teacher_id, w.discipline, w.semester,
                          w.total_hours, w.lesson_type, g.name as group_name, t.full_name as teacher_name
                   FROM workload w
                   JOIN groups g ON w.group_id = g.id
                   JOIN teachers t ON w.teacher_id = t.id''')
    cols = [d[0] for d in cur.description]
    workloads = [dict(zip(cols, r)) for r in cur.fetchall()]

    cur.execute('SELECT id, name FROM academic_years')
    academic_years = {r[1]: r[0] for r in cur.fetchall()}

    return groups, teachers, rooms, workloads, academic_years


def resolve_room(room_text, rooms_db):
    if not room_text:
        return None
    room_text = str(room_text).strip()
    if room_text.endswith('.0'):
        room_text = room_text[:-2]

    for key, val in ROOM_SPECIAL_MAP.items():
        if room_text.lower() == key.lower():
            return rooms_db.get(val)

    if room_text in rooms_db:
        return rooms_db[room_text]

    try:
        num = str(int(float(room_text)))
        if num in rooms_db:
            return rooms_db[num]
    except (ValueError, TypeError):
        pass
    return None


def resolve_teacher(teacher_text, teachers_db):
    if not teacher_text:
        return None
    teacher_text = str(teacher_text).strip()
    if teacher_text in teachers_db:
        return teachers_db[teacher_text]
    if teacher_text in TEACHER_NORMALIZE:
        normalized = TEACHER_NORMALIZE[teacher_text]
        if normalized in teachers_db:
            return teachers_db[normalized]
    surname = teacher_text.split()[0] if teacher_text else ''
    candidates = [name for name in teachers_db if name.startswith(surname)]
    if len(candidates) == 1:
        return teachers_db[candidates[0]]
    return None


def find_workload(group_name, discipline_text, semester, workloads):
    if not discipline_text:
        return None
    discipline_text = str(discipline_text).strip()

    group_workloads = [w for w in workloads
                       if w['group_name'] == group_name and w['semester'] == semester]
    if not group_workloads:
        return None

    best_match = None
    best_score = 0

    for w in group_workloads:
        db_disc = w['discipline']
        compare_len = min(len(discipline_text), len(db_disc))
        if compare_len < 5:
            continue
        match_len = 0
        for i in range(compare_len):
            if discipline_text[i] == db_disc[i]:
                match_len += 1
            else:
                break
        if match_len > best_score and match_len >= 10:
            best_score = match_len
            best_match = w

    return best_match


def parse_excel_sheet(ws, config, groups_db):
    header_row = config['header_row']
    pair_time_map = config['pair_time_map']

    group_columns = {}
    for cell in ws[header_row]:
        if cell.value and cell.column >= 4 and cell.column % 2 == 0:
            group_name = str(cell.value).strip()
            if group_name in groups_db and '.' not in group_name:
                group_columns[cell.column] = group_name

    print(f"  Found {len(group_columns)} groups: {list(group_columns.values())}")

    entries = []
    current_day = None
    current_day_num = None
    row_num = config['data_start_row']

    while row_num <= ws.max_row:
        row_cells = {}
        for cell in ws[row_num]:
            if cell.value is not None:
                row_cells[cell.column] = cell.value

        if not row_cells:
            row_num += 1
            continue

        if 1 in row_cells:
            day_text = str(row_cells[1]).strip().replace('\n', '')
            if day_text.upper() in DAY_NAMES:
                current_day = day_text.upper()
                current_day_num = DAY_NAMES[current_day]

        if 3 in row_cells and current_day is not None:
            pair_roman = str(row_cells[3]).strip()
            if pair_roman in pair_time_map:
                pair_number = pair_time_map[pair_roman]
                disc_row = row_cells

                teacher_row = {}
                if row_num + 1 <= ws.max_row:
                    for cell in ws[row_num + 1]:
                        if cell.value is not None:
                            teacher_row[cell.column] = cell.value

                for col, group_name in group_columns.items():
                    disc_text = disc_row.get(col)
                    if not disc_text:
                        continue
                    disc_text = str(disc_text).strip()
                    if not disc_text:
                        continue

                    room_hint = disc_row.get(col + 1)
                    teacher_text = teacher_row.get(col)
                    teacher_text = str(teacher_text).strip() if teacher_text else None
                    room_actual = teacher_row.get(col + 1)

                    entries.append({
                        'day_num': current_day_num,
                        'pair_number': pair_number,
                        'group_name': group_name,
                        'discipline': disc_text,
                        'teacher': teacher_text,
                        'room_hint': str(room_hint).strip() if room_hint else None,
                        'room_actual': room_actual,
                    })

                row_num += 2
                continue

        row_num += 1

    return entries


def get_weeks(start_date, end_date):
    weeks = []
    current = start_date
    while current <= end_date:
        monday = current - timedelta(days=current.weekday())
        if monday not in weeks:
            weeks.append(monday)
        current += timedelta(days=7)
    return weeks


def main():
    conn = sqlite3.connect(DB_PATH)
    groups_db, teachers_db, rooms_db, workloads, academic_years = load_db_data(conn)
    cur = conn.cursor()

    ay_id = None
    for name, aid in academic_years.items():
        if '2025' in name or '2026' in name:
            ay_id = aid
            break
    if not ay_id:
        ay_id = list(academic_years.values())[0] if academic_years else 1
    print(f"Academic year ID: {ay_id}")

    weeks = get_weeks(DATE_START, DATE_END)
    print(f"Weeks: {[w.strftime('%d.%m.%Y') for w in weeks]}")

    total_schedule = 0
    total_templates = 0
    total_unmatched = 0
    unmatched_list = []

    for fc in EXCEL_FILES:
        print(f"\n{'='*60}")
        print(f"File: {fc['path']}")
        print(f"{'='*60}")

        wb = openpyxl.load_workbook(fc['path'], data_only=True)

        ws_tpl = wb[wb.sheetnames[fc['sheet_index']]]
        print(f"\nTemplate sheet: {wb.sheetnames[fc['sheet_index']]}")
        tpl_entries = parse_excel_sheet(ws_tpl, fc, groups_db)
        print(f"  {len(tpl_entries)} slots parsed")

        ws_w1 = wb[wb.sheetnames[fc['sheet_index_week1']]]
        print(f"Week1 sheet: {wb.sheetnames[fc['sheet_index_week1']]}")
        w1_entries = parse_excel_sheet(ws_w1, fc, groups_db)
        print(f"  {len(w1_entries)} slots parsed")

        wb.close()

        semester = fc['semester']
        wl_pattern = {}

        def process_entries(entries, week_monday, label):
            nonlocal total_schedule, total_unmatched
            created = 0
            unmatched = 0

            for e in entries:
                wl = find_workload(e['group_name'], e['discipline'], semester, workloads)
                if not wl:
                    unmatched += 1
                    unmatched_list.append(
                        f"  [{label}] {e['group_name']} | {e['discipline'][:50]} | t={e['teacher']}")
                    continue

                wl_id = wl['id']
                grp_id = wl['group_id']

                room_id = None
                if e['room_actual']:
                    room_id = resolve_room(e['room_actual'], rooms_db)
                if not room_id and e['room_hint']:
                    room_id = resolve_room(e['room_hint'], rooms_db)

                date = week_monday + timedelta(days=e['day_num'])
                date_str = date.strftime('%Y-%m-%d')

                cur.execute('SELECT id FROM schedule WHERE workload_id=? AND date=? AND pair_number=?',
                           (wl_id, date_str, e['pair_number']))
                if cur.fetchone():
                    continue

                now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                cur.execute('''INSERT INTO schedule (workload_id, room_id, date, pair_number, status, created_at, updated_at)
                              VALUES (?, ?, ?, ?, 'planned', ?, ?)''',
                           (wl_id, room_id, date_str, e['pair_number'], now, now))
                created += 1

                if wl_id not in wl_pattern:
                    wl_pattern[wl_id] = {
                        'group_id': grp_id,
                        'group_name': e['group_name'],
                        'discipline': wl['discipline'],
                        'slots': {},
                        'rooms': {}
                    }
                if e['day_num'] not in wl_pattern[wl_id]['slots']:
                    wl_pattern[wl_id]['slots'][e['day_num']] = []
                if e['pair_number'] not in wl_pattern[wl_id]['slots'][e['day_num']]:
                    wl_pattern[wl_id]['slots'][e['day_num']].append(e['pair_number'])
                    wl_pattern[wl_id]['rooms'][(e['day_num'], e['pair_number'])] = room_id

            total_schedule += created
            total_unmatched += unmatched
            print(f"  [{label}] +{created} schedule, {unmatched} unmatched")

        if weeks:
            process_entries(w1_entries, weeks[0], f"W1 {weeks[0].strftime('%d.%m')}")
        for wm in weeks[1:]:
            process_entries(tpl_entries, wm, f"Tpl {wm.strftime('%d.%m')}")

        conn.commit()

        # Create WorkloadTemplates
        print(f"\nCreating templates...")
        day_names_ru = {0: 'Понедельник', 1: 'Вторник', 2: 'Среда',
                        3: 'Четверг', 4: 'Пятница', 5: 'Суббота'}

        for wl_id, pat in wl_pattern.items():
            grp_id = pat['group_id']
            distribution = []
            week_data = {'week': 1, 'hours': 0, 'schedule': []}

            for day_num in sorted(pat['slots'].keys()):
                day_sched = {
                    'day': day_names_ru.get(day_num, str(day_num)),
                    'day_num': day_num,
                    'pairs': []
                }
                for pair in sorted(pat['slots'][day_num]):
                    rid = pat['rooms'].get((day_num, pair))
                    day_sched['pairs'].append({'pair': pair, 'status': 'selected', 'room_id': rid})
                    week_data['hours'] += 2
                day_sched['pairs_count'] = len(day_sched['pairs'])
                week_data['schedule'].append(day_sched)

            distribution.append(week_data)
            dist_json = json.dumps(distribution, ensure_ascii=False)

            cur.execute('SELECT id FROM workload_templates WHERE workload_id=? AND group_id=? AND semester=?',
                       (wl_id, grp_id, semester))
            existing = cur.fetchone()

            tpl_name = f"Excel {pat['group_name']} - {pat['discipline'][:50]}"
            total_hrs = week_data['hours']
            now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

            if existing:
                cur.execute('''UPDATE workload_templates SET distribution_json=?, name=?, total_hours=?, updated_at=?
                              WHERE id=?''', (dist_json, tpl_name, total_hrs, now, existing[0]))
            else:
                cur.execute('''INSERT INTO workload_templates
                              (name, workload_id, group_id, academic_year_id, semester, total_hours, distribution_json, created_at, created_by)
                              VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                           (tpl_name, wl_id, grp_id, ay_id, semester, total_hrs, dist_json, now, 'excel_import'))
                total_templates += 1

        conn.commit()
        print(f"  Templates: {len(wl_pattern)} processed")

    # Summary
    print(f"\n{'='*60}")
    print(f"DONE")
    print(f"{'='*60}")
    print(f"Schedule created: {total_schedule}")
    print(f"Templates created: {total_templates}")
    print(f"Unmatched: {total_unmatched}")

    if unmatched_list:
        print(f"\nUnmatched (first 40):")
        for item in unmatched_list[:40]:
            print(item)

    cur.execute('''SELECT COUNT(*) FROM schedule s JOIN workload w ON s.workload_id=w.id
                   JOIN groups g ON w.group_id=g.id WHERE g.enrollment_year IN (2022, 2024)''')
    print(f"\nVerify - total schedule for courses 2+4: {cur.fetchone()[0]}")

    conn.close()


if __name__ == '__main__':
    main()
