#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Update ONLY room assignments from Excel
Does NOT change pairs, teachers, or schedules - ONLY rooms!
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.path.insert(0, '.')

import openpyxl
from datetime import date, timedelta
from app import create_app, db
from app.models import Schedule, Workload, Group, Room

app = create_app()

EXCEL_FILE = 'Копия 1 курс Расписание 2 семестр (1).xlsx'
SHEET_NAME = 'II cем'

# Group columns mapping: group_col -> room_col
GROUP_COLS = {
    4: ('БНГС 1-25', 5),
    6: ('ТДНГ 1-25', 7),
    8: ('ЭНГМ 1-25', 9),
    10: ('ЭС 1-25', 11),
    12: ('ЭС 2-25', 13),
    14: ('ЭС 3-25', 15),
    16: ('ХТП 1-25', 17),
    18: ('ХТП 2-25', 19),
    20: ('ХТП 3-25', 21),
    22: ('АиУ 1-25', 23),
    24: ('АиУ 2-25', 25),
    26: ('АиУ 3-25', 27),
    28: ('АиУ 4-25', 29),
    30: ('АиУ 5-25', 31),
    32: ('ПО 1-25', 33),
    34: ('ПО 2-25', 35),
}

# Roman numerals to numbers
ROMAN_MAP = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6}

# Day mapping
DAY_MAP = {
    'ПОНЕДЕЛЬНИК': 0,
    'ВТОРНИК': 1,
    'СРЕДА': 2,
    'ЧЕТВЕРГ': 3,
    'ПЯТНИЦА': 4,
}

def get_or_create_room(room_number):
    """Get room by number or create if not exists"""
    if not room_number:
        return None

    room_number = str(room_number).strip()
    if not room_number or room_number == '-' or room_number == 'None':
        return None

    # Clean up room number
    room_number = room_number.replace('.0', '')  # Remove .0 from floats

    room = Room.query.filter_by(number=room_number).first()
    if not room:
        room = Room(number=room_number, room_type='учебная')
        db.session.add(room)
        db.session.flush()
        print(f'  Created new room: {room_number}')

    return room

def parse_room_from_excel(room_cell_value):
    """
    Parse room from Excel cell
    Format can be: "105", "209/207", "С/з/208", "105.0"
    Returns: (left_room, right_room) or (single_room, single_room)
    """
    if not room_cell_value:
        return None, None

    room_str = str(room_cell_value).strip()
    if not room_str or room_str == '-' or room_str == 'None':
        return None, None

    # Handle special cases
    room_str = room_str.replace('.0', '')  # Remove .0

    # Check if it's a split room (left/right)
    if '/' in room_str:
        parts = room_str.split('/')
        if len(parts) == 2:
            return parts[0].strip(), parts[1].strip()
        elif len(parts) == 3:
            # Could be "С/з/208" (спортзал) or "202/203/105"
            # If first part is "С" it's спортзал
            if parts[0] == 'С' and parts[1] == 'з':
                return 'С/з', parts[2].strip()
            else:
                # Three rooms - take first two as left/right
                return parts[0].strip(), parts[1].strip()
        else:
            # Multiple rooms, take first as default
            return parts[0].strip(), parts[0].strip()

    # Single room for both weeks
    return room_str, room_str

def parse_excel_rooms():
    """Parse room assignments from Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    # Structure: {group_name: {day_name: {pair_num: {'left_room': X, 'right_room': Y}}}}
    room_data = {}

    current_day = None
    current_pair = None

    for row_idx in range(6, ws.max_row + 1):
        # Check for day name in column 1
        col1 = ws.cell(row=row_idx, column=1).value
        if col1:
            col1_str = str(col1).strip().upper().replace('\n', '')
            for day_name in DAY_MAP.keys():
                if day_name in col1_str:
                    current_day = day_name
                    break

        # Check for pair number in column 3
        col3 = ws.cell(row=row_idx, column=3).value
        if col3:
            col3_str = str(col3).strip()
            if col3_str in ROMAN_MAP:
                current_pair = ROMAN_MAP[col3_str]

        if not current_day or not current_pair:
            continue

        # Read rooms for each group
        for group_col, (group_name, room_col) in GROUP_COLS.items():
            room_value = ws.cell(row=row_idx, column=room_col).value

            if room_value:
                left_room, right_room = parse_room_from_excel(room_value)

                if left_room or right_room:
                    if group_name not in room_data:
                        room_data[group_name] = {}
                    if current_day not in room_data[group_name]:
                        room_data[group_name][current_day] = {}

                    room_data[group_name][current_day][current_pair] = {
                        'left_room': left_room,
                        'right_room': right_room
                    }

    return room_data

# Week calculation
ACADEMIC_YEAR_START = date(2025, 9, 1)

def get_week_type(d):
    """Get week type: LEFT (odd) or RIGHT (even)"""
    days_since_monday = ACADEMIC_YEAR_START.weekday()
    week_start = ACADEMIC_YEAR_START - timedelta(days=days_since_monday)
    days_diff = (d - week_start).days
    week_num = (days_diff // 7) + 1
    return 'RIGHT' if week_num % 2 == 0 else 'LEFT'

def update_rooms(dry_run=True):
    """Update room assignments in database"""

    print("=" * 60)
    print("ROOM UPDATE FROM EXCEL")
    print("=" * 60)

    # Parse Excel
    print("\n1. Parsing rooms from Excel...")
    room_data = parse_excel_rooms()

    groups_with_rooms = len(room_data)
    print(f"   Found room data for {groups_with_rooms} groups")

    # Show sample
    for group in list(room_data.keys())[:2]:
        print(f"\n   {group}:")
        for day in list(room_data[group].keys())[:2]:
            for pair, rooms in room_data[group][day].items():
                print(f"      {day} Para {pair}: L={rooms['left_room']}, R={rooms['right_room']}")

    if dry_run:
        print("\n[DRY RUN - no changes will be made]")

    # Update database
    print("\n2. Updating schedules...")

    with app.app_context():
        updated = 0
        not_found = 0

        # Get all schedules for semester 2
        start_date = date(2026, 1, 12)
        end_date = date(2026, 6, 30)

        for group_name, days in room_data.items():
            group = Group.query.filter_by(name=group_name).first()
            if not group:
                print(f"   Warning: Group {group_name} not found")
                continue

            for day_name, pairs in days.items():
                weekday = DAY_MAP[day_name]

                for pair_num, rooms in pairs.items():
                    left_room_num = rooms['left_room']
                    right_room_num = rooms['right_room']

                    # Find all schedules for this group/day/pair
                    schedules = Schedule.query.join(Workload).filter(
                        Workload.group_id == group.id,
                        Schedule.date >= start_date,
                        Schedule.date <= end_date,
                        Schedule.pair_number == pair_num
                    ).all()

                    for schedule in schedules:
                        if schedule.date.weekday() != weekday:
                            continue

                        # Determine which room to use based on week type
                        week_type = get_week_type(schedule.date)
                        room_num = left_room_num if week_type == 'LEFT' else right_room_num

                        if room_num:
                            if not dry_run:
                                room = get_or_create_room(room_num)
                                if room and schedule.room_id != room.id:
                                    schedule.room_id = room.id
                                    updated += 1
                            else:
                                updated += 1

        if not dry_run:
            db.session.commit()

        print(f"\n   Updated {updated} schedule entries with rooms")

    return updated

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--apply', action='store_true', help='Apply changes')
    args = parser.parse_args()

    update_rooms(dry_run=not args.apply)
